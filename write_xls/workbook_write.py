"""
Author: Fedyaev I.

class WorkBookContractWells write Excel data contracts statistics
"""
# openpyxl
# class
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl.styles import Color, PatternFill, Font, Border, Alignment, Side
# constant
from openpyxl.styles.borders import BORDER_THIN, BORDER_MEDIUM
# colors
from openpyxl.styles.colors import WHITE
from alexey_fedyaev.const.const import *
# utils
from openpyxl.utils import get_column_letter
#
from alexey_fedyaev.data.container import BorderStatistWells
from alexey_fedyaev.utils.utils import terminal_progress_bar

# array month of year
g_arr_month_by_year = ["Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
                       "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"]
# help columns
g_arr_util = ["Единица измерения", "Декабрь 2019", "ИТОГО 2019"]


class WorkBookContractWells:
    """
    write data contracts to Excel
    """
    NAME_FONT = "Calibri"
    PATTERN_TYPE = "solid"
    ALIGNMENT_VH = "center"
    # border
    S_BORDER_NORM = BORDER_THIN
    S_BORDER_HIGH = BORDER_MEDIUM
    BORDER_NORM = Side(style=S_BORDER_NORM)
    BORDER_HIGH = Side(style=S_BORDER_HIGH)

    def __init__(self):
        self.work_book = Workbook()
        self.work_book.remove(self.work_book.active)

    def create_sheet(self, title=None, index=None) -> Worksheet:
        """
        create sheet and return object sheet
        :param title: title sheet
        :param index: index set sheet
        :return: object new sheet (WorkSheet)
        """
        return self.work_book.create_sheet(title=title, index=index)

    @staticmethod
    def get_cell_name(sheet: Worksheet,
                      row: int, col: int) -> str:
        """
        get cell name "A1" ...
        :param sheet: sheet get name
        :param row: row index
        :param col: column index
        :return: string name cell
        """
        return sheet.cell(row=row + 1, column=col + 1).coordinate

    @staticmethod
    def get_cell(sheet: Worksheet,
                 row: int, col: int) -> Cell:
        """
        get cell data
        :param sheet: sheet get info cell
        :param row: index row
        :param col: index col
        :return: cell info data
        """
        return sheet.cell(row=row + 1, column=col + 1)

    @staticmethod
    def set_cell_value(sheet: Worksheet,
                       row: int, column: int,
                       val,
                       color: str = WHITE,
                       is_border: bool = False, is_bold: bool = False, is_italic: bool = False) -> None:
        """
        set value to cell
        :param sheet: sheet data
        :param row: index row
        :param column: index col
        :param val: value
        :param color: [if need] color
        :param is_border: [if need] border
        :param is_bold:  [if need] bold flag
        :param is_italic: [if need] italic flag
        :return:
        """
        d: Cell = sheet.cell(row=row + 1, column=column + 1, value=val)
        fgColor = Color(rgb=color)
        d.fill = PatternFill(patternType=WorkBookContractWells.PATTERN_TYPE,
                             fgColor=fgColor)
        d.alignment = Alignment(vertical=WorkBookContractWells.ALIGNMENT_VH,
                                horizontal=WorkBookContractWells.ALIGNMENT_VH)
        if is_border:
            _style = WorkBookContractWells.BORDER_NORM
            d.border = Border(left=_style, right=_style,
                              top=_style, bottom=_style)
        d.font = Font(name=WorkBookContractWells.NAME_FONT,
                      bold=is_bold, italic=is_italic)
        return

    def set_row_color_bold(self, sheet: Worksheet,
                           row: int,
                           beg_col: int, end_col: int, is_bold: bool = True, color: str = GREEN):
        """
        set green row by border column
        :param sheet: sheet object
        :param row: index row
        :param beg_col: begin index column
        :param end_col: end index column
        :param is_bold: bold flag
        :param color: color
        :return:
        """
        fgColor = Color(rgb=color)
        for col in range(beg_col, end_col + 1, 1):
            d: Cell = self.get_cell(sheet=sheet, row=row, col=col)
            d.font = Font(name=WorkBookContractWells.NAME_FONT,
                          bold=is_bold)
            d.fill = PatternFill(patternType=WorkBookContractWells.PATTERN_TYPE,
                                 fgColor=fgColor)
        return

    def set_grid_region_border(self, sheet: Worksheet,
                               beg_row: int, end_row: int,
                               beg_col: int, end_col: int) -> None:
        """
        set region border grid
        :param sheet: sheet object
        :param beg_row: begin index row
        :param end_row: end index row
        :param beg_col: begon index column
        :param end_col: end index column
        :return:
        """
        # style
        _style = WorkBookContractWells.BORDER_NORM
        _style_high = WorkBookContractWells.BORDER_HIGH
        # set all border
        for row in range(beg_row, end_row + 1, 1):
            for col in range(beg_col, end_col + 1, 1):
                d: Cell = self.get_cell(sheet=sheet, row=row, col=col)
                d.border = Border(left=_style, right=_style, top=_style, bottom=_style)
        # angle
        # up|down
        for col in range(beg_col, end_col + 1, 1):
            # up
            d: Cell = self.get_cell(sheet=sheet, row=beg_row, col=col)
            d.border = Border(top=_style_high, bottom=_style, left=_style, right=_style)
            # down
            d: Cell = self.get_cell(sheet=sheet, row=end_row, col=col)
            d.border = Border(top=_style, bottom=_style_high, left=_style, right=_style)
        # left | right
        for row in range(beg_row, end_row + 1, 1):
            # left
            d: Cell = self.get_cell(sheet=sheet, row=row, col=beg_col)
            d.border = Border(top=_style, bottom=_style, left=_style_high, right=_style)
            # right
            d: Cell = self.get_cell(sheet=sheet, row=row, col=end_col)
            d.border = Border(top=_style, bottom=_style, left=_style, right=_style_high)
        # up left
        d: Cell = self.get_cell(sheet=sheet, row=beg_row, col=beg_col)
        d.border = Border(left=_style_high, top=_style_high, bottom=_style, right=_style)
        # down left
        d: Cell = self.get_cell(sheet=sheet, row=end_row, col=beg_col)
        d.border = Border(left=_style_high, bottom=_style_high, top=_style, right=_style)
        # right up
        d: Cell = self.get_cell(sheet=sheet, row=beg_row, col=end_col)
        d.border = Border(left=_style, top=_style_high, bottom=_style, right=_style_high)
        # right down
        d: Cell = self.get_cell(sheet=sheet, row=end_row, col=end_col)
        d.border = Border(left=_style, top=_style, bottom=_style_high, right=_style_high)
        return

    @staticmethod
    def merge_cells(sheet: Worksheet,
                    beg_row: int, end_row: int,
                    beg_col: int, end_col: int) -> None:
        """
        merge cells by index
        :param sheet: sheet object
        :param beg_row: begin index row
        :param end_row: end index row
        :param beg_col: begin index column
        :param end_col: end index column
        :return:
        """
        sheet.merge_cells(start_row=beg_row + 1, end_row=end_row + 1,
                          start_column=beg_col + 1, end_column=end_col + 1)
        return

    def write_to_sheet(self, beg_row: int, beg_col: int,
                       sheet: Worksheet,
                       lb: float, ub: float,
                       all_border_statistic_wells: dict) -> int:
        """
        write sheet data border
        :param beg_row: begin index row
        :param beg_col: begin index column
        :param sheet: sheet object
        :param lb: lower bound border
        :param ub: upper bound border
        :param all_border_statistic_wells: border statistic array by month
        :return: end index row after fill table
        """
        # create data write
        arr_data_wells = [{}, {}, {}, {}]
        for key_month in all_border_statistic_wells.keys():
            item: BorderStatistWells = all_border_statistic_wells[key_month]
            #
            arr_data_wells[0][key_month] = item.count_wells
            arr_data_wells[1][key_month] = item.count_days
            arr_data_wells[2][key_month] = item.costs_day
            arr_data_wells[3][key_month] = item.sum_costs
        # heap data
        self.merge_cells(sheet=sheet,
                         beg_row=beg_row, end_row=beg_row,
                         beg_col=beg_col, end_col=beg_col + 1)
        s_text_lub = ""
        if lb == ub:
            self.set_cell_value(sheet=sheet,
                                row=beg_row, column=beg_col,
                                val="УЭЦН {0} м3/сут.".format(int(lb)),
                                color=YELLOW, is_border=True)
            s_text_lub = "bound {0:7.1f}".format(lb)
        else:
            self.set_cell_value(sheet=sheet,
                                row=beg_row, column=beg_col,
                                val="УЭЦН от {0} м3/сут. до {1} м3.сут.".format(int(lb), int(ub)),
                                color=YELLOW, is_border=True)
            s_text_lub = "bound {0:7.1f} {1:7.1f}".format(lb, ub)

        #
        _cur_beg_col = beg_col + 2
        for item in [*g_arr_util, *g_arr_month_by_year]:
            self.set_cell_value(sheet=sheet,
                                row=beg_row, column=_cur_beg_col,
                                val=item,
                                color=YELLOW, is_border=True)
            _cur_beg_col += 1
        # set data
        _arr_heap = ["Фонд", "Количество суток", "Стоимость одних суток", "Общие затраты по группе"]
        _arr_unit = ["скважина", "сутки", "руб./сутки", "руб."]
        beg_row += 1
        # value column
        _val_beg_col = beg_col + 2
        _val_end_col = _val_beg_col + len(g_arr_util) + len(g_arr_month_by_year) - 1
        #
        _idx_step_fact = 5
        _n_all_step = len(arr_data_wells) * len(arr_data_wells[0].keys())
        step_pct = 100 / _n_all_step
        count = 0
        #
        for s_heap, s_unit, arr_data_month in zip(_arr_heap, _arr_unit, arr_data_wells):
            # set left heap
            self._write_left_heap(beg_row=beg_row, beg_col=beg_col,
                                  sheet=sheet,
                                  s_unit=s_unit, s_heap=s_heap)
            # set border
            self.set_grid_region_border(sheet=sheet,
                                        beg_row=beg_row, end_row=beg_row + 10,
                                        beg_col=_val_beg_col + 1,
                                        end_col=_val_end_col)
            # loop set value month
            for key_month in arr_data_month:
                # progerss_bar in tarminal
                terminal_progress_bar(pct=(count + 1) * step_pct, text=s_text_lub)
                count += 1
                #
                cur_col = _val_beg_col + len(g_arr_util) + (key_month - 1)
                cur_row = beg_row + _idx_step_fact
                self.set_cell_value(sheet=sheet,
                                    row=cur_row, column=cur_col,
                                    val=arr_data_month[key_month], is_bold=True)
                s_name = self.get_cell_name(sheet=sheet, row=cur_row, col=cur_col)
                #
                self.set_cell_value(sheet=sheet,
                                    row=cur_row - 1, column=cur_col,
                                    val="={0}".format(s_name))
                #
                self.set_cell_value(sheet=sheet,
                                    row=cur_row - 3, column=cur_col,
                                    val="={0}".format(s_name))
                _max_step_row = 5
                #
                # loop fill formula data
                for step_row in range(1, _max_step_row + 1, 1):
                    s_name_up = self.get_cell_name(sheet=sheet, row=cur_row - _max_step_row + step_row - 1,
                                                   col=cur_col)
                    self.set_cell_value(sheet=sheet,
                                        row=cur_row + step_row,
                                        column=cur_col,
                                        val="={0}-{1}".format(s_name, s_name_up))
                # end loop fill formula data
            # end loop set value month
            # set green
            self.set_row_color_bold(sheet=sheet, row=beg_row + _idx_step_fact,
                                    beg_col=_val_beg_col - 1, end_col=_val_end_col)
            # go next
            beg_row += 11
        print()
        return beg_row

    def _write_left_heap(self,
                         beg_row: int, beg_col: int,
                         sheet: Worksheet,
                         s_unit: str, s_heap: str) -> int:
        """
        write left heap
        :param beg_row: begin index row
        :param beg_col: begin index column
        :param sheet: sheet object
        :param s_unit: string unit
        :param s_heap: string heap
        :return: end fill index row
        """
        # left merge heap
        self.merge_cells(sheet=sheet,
                         beg_row=beg_row, end_row=beg_row + 10,
                         beg_col=beg_col, end_col=beg_col)
        self.set_cell_value(sheet=sheet,
                            row=beg_row, column=beg_col,
                            val=s_heap, is_border=True)
        #
        # set units
        beg_col += 1
        _arr_bp = ["БП-0", "ПП", "БП-1", "БП-2", "...", ]
        for item in [*_arr_bp, "факт", *["отклонение от {0}".format(item) for item in _arr_bp]]:
            #
            self.set_cell_value(sheet, row=beg_row, column=beg_col, val=item, is_border=True)
            # units
            self.set_cell_value(sheet, row=beg_row, column=beg_col + 1, val=s_unit, is_border=True)
            beg_row += 1
        return beg_row

    @staticmethod
    def auto_format_cell_width(sheet: Worksheet) -> None:
        """
        set auto width cell by column
        :param sheet: sheet object
        :return:
        """
        for col in sheet.columns:
            max_length = 0
            column = col[0].column  # Get the column name
            for cell in col:
                if cell.coordinate in sheet.merged_cells:  # not check merge_cells
                    continue
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[get_column_letter(column)].width = adjusted_width
        sheet.column_dimensions[get_column_letter(1)].width = 30
        sheet.column_dimensions[get_column_letter(2)].width = 40
        return

    def save(self, path: str) -> None:
        """
        save Excel by path
        :param path: full path save Excel
        :return:
        """
        self.work_book.save(filename=path)
        return
