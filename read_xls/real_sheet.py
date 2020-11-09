# pandas
import pandas as pd
# openpyxl
# read
from openpyxl import load_workbook
# const
from const.const import *


def read_list_sheets(path_xls: str):
    """
    read input sheets by lists
    :param path_xls: full path Excel
    :return: list sheets
    """
    return load_workbook(path_xls, read_only=True).sheetnames


def extract_xls(path_xls: str, name_sheet: str) -> pd.DataFrame:  #
    """
    extract life input data in sheet
    :param path_xls: full path Excel
    :return: data Frame sheets
    """
    df = pd.read_excel(path_xls, sheet_name=name_sheet,
                       names=[
                           g_key_number, g_key_contract, g_key_NGDU,
                           g_key_field, g_key_well_number, g_key_bush_number,
                           g_key_liquid, g_key_count_day, g_key_data_install,
                           g_key_data_uninstall, g_key_reason_stop, g_key_type_YECN,
                           g_key_owner, g_key_sum_tex_close, g_key_TK,
                           g_key_EE, g_key_fund, g_key_move_fund,
                           g_key_refusal_CNO, g_key_refusal_MRP, g_key_sum_day_rent,
                           g_key_type_stop
                       ], skiprows=2)
    #
    arr_na_key_fund = df[g_key_fund].notna()
    count_dead = 0
    for i in range(len(arr_na_key_fund) - 1, 0, -1):
        if arr_na_key_fund[i] == False:
            count_dead += 1
        else:
            break
    # fill true all
    for i in range(0, len(arr_na_key_fund) - count_dead, 1):
        arr_na_key_fund[i] = True
    # filter
    df = df[arr_na_key_fund]
    return df
